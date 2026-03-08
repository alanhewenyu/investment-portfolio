#!/usr/bin/env python3
"""Import portfolio data from investment_2026.xlsm into SQLite."""

import os
import sys
import openpyxl
from db import init_db, get_conn, upsert_position, insert_closed_trade, upsert_cash, upsert_nav, upsert_fx, upsert_margin

EXCEL_PATH = os.environ.get(
    'PORTFOLIO_EXCEL',
    os.path.expanduser('~/Desktop/investment_2026.xlsm'),
)

# ── Name → Yahoo Finance ticker mapping ──────────────────
# Excel column A has formulas (#VALUE!) for 港股/美股/基金.
# We build a name→ticker map and fall back to 股价数据源 sheet.

_NAME_TO_TICKER = {
    # 港股
    '美团': '3690.HK',
    '古茗': '1364.HK',
    '顺丰': '6936.HK',
    '安踏': '2020.HK',
    '太古A': '0019.HK',
    '小米': '1810.HK',
    '巨子生物': '2367.HK',
    '阅文集团': '0772.HK',
    '大麦娱乐': '1060.HK',
    '腾讯控股': '0700.HK',
    '美的集团': '0300.HK',
    '招商银行': '3968.HK',
    '中烟香港': '6055.HK',
    '汇丰控股': '0005.HK',
    '普拉达': '1913.HK',
    '中金公司': '3908.HK',
    '香港交易所': '0388.HK',
    '中国海洋石油': '0883.HK',
    '蜜雪': '2097.HK',
    # 美股
    '谷歌': 'GOOGL',
    '英伟达': 'NVDA',
    '特斯拉': 'TSLA',
    '台积电': 'TSM',
    '英特尔': 'INTC',
    '星座': 'STZ',
    '20+美国国债': 'TLT',
    '富途': 'FUTU',
    '苹果': 'AAPL',
    '拉夫劳伦': 'RL',
    '纳指100': 'QQQ',
    '标普500': 'SPY',
    '中国大盘股ETF': 'FXI',
    'Circle': 'CRCL',
    '滴滴': 'DIDIY',
    '腾讯音乐': 'TME',
    # 基金 (no Yahoo ticker; use fund code)
    '华安媒体互联网': '001071',
    '汇添富美丽30': '000173',
}

# B股 currency keywords
_CURRENCY_MAP = {'美金': 'USD', '港币': 'HKD', 'USD': 'USD', 'HKD': 'HKD'}


def _build_name_ticker_map(wb):
    """Build name→ticker from 股价数据源 sheet to supplement _NAME_TO_TICKER."""
    mapping = dict(_NAME_TO_TICKER)
    if '股价数据源' not in wb.sheetnames:
        return mapping
    ws = wb['股价数据源']
    for row in range(2, ws.max_row + 1):
        ticker = ws[f'A{row}'].value
        name = ws[f'B{row}'].value
        if ticker and name and isinstance(ticker, str) and isinstance(name, str):
            name = name.strip()
            # Remove trailing "B股"/"B" suffix for matching
            short_name = name.replace('B股', '').replace('B', '').strip()
            mapping[name] = ticker.strip()
            if short_name != name:
                mapping[short_name] = ticker.strip()
    return mapping


def _resolve_ticker(name, col_a_value, market, name_map):
    """Resolve Yahoo Finance ticker for a position.

    Priority:
    1. Name→ticker mapping from 股价数据源 (most reliable)
    2. Column A value if it's a valid ticker (not #VALUE! or formula)
    3. Fallback: name itself
    """
    # Try name mapping first (股价数据源 is authoritative)
    clean_name = name.strip() if name else ''
    if clean_name in name_map:
        return name_map[clean_name]
    # Try stripping B股 suffix for B-share matching
    for suffix in ('B股', 'B'):
        if clean_name.endswith(suffix):
            short = clean_name[:-len(suffix)]
            if short in name_map:
                return name_map[short]

    # Try column A (works for A股 mainly)
    if col_a_value and isinstance(col_a_value, str):
        raw = col_a_value.strip()
        if raw and '#' not in raw and '=' not in raw and '合计' not in raw:
            # Normalize: .SH → .SS for Shanghai; .t → .T for Japan
            if raw.endswith('.SH'):
                return raw.replace('.SH', '.SS')
            if raw.endswith('.t'):
                return raw[:-2] + '.T'
            return raw

    return clean_name  # fallback


# ── Sheet parsers ─────────────────────────────────────────

_SKIP_KEYWORDS = ('合计', '场内杠杆', '已清仓', '净资产', '自有资金',
                   '累计', '期末', '当年', '投入资金', '证券市值',
                   '小计', '名称', '清仓盈亏明细')


def _is_skip_row(text):
    if not text or not isinstance(text, str):
        return True
    return any(k in text for k in _SKIP_KEYWORDS)


def _parse_position_sheet(ws, market, default_currency, default_broker, name_map):
    """Parse a standard position sheet.

    Active positions: row has qty != 0 and cost != None
    Closed trades section: after '清仓盈亏明细' marker, format: A=name, J=pnl
    """
    positions = []
    closed = []
    in_closed_section = False

    for row in range(5, ws.max_row + 1):
        col_a = ws[f'A{row}'].value
        name = ws[f'B{row}'].value

        # Detect closed trades section marker
        marker = str(col_a or '') + str(name or '')
        if '清仓盈亏明细' in marker:
            in_closed_section = True
            continue

        if in_closed_section:
            # Closed trade format: A=name, C=broker (optional), J=P&L
            trade_name = str(col_a or '').strip() if col_a else ''
            if not trade_name or _is_skip_row(trade_name):
                continue
            pnl = ws[f'J{row}'].value
            if pnl is not None and isinstance(pnl, (int, float)) and pnl != 0:
                broker = ws[f'C{row}'].value
                broker = str(broker).strip() if broker else default_broker
                closed.append({
                    'ticker': name_map.get(trade_name, trade_name),
                    'name': trade_name,
                    'market': market,
                    'broker': broker,
                    'currency': default_currency,
                    'realized_pnl': float(pnl),
                })
            continue

        # Active positions section
        if not name or not isinstance(name, str):
            continue
        name = str(name).strip()
        if _is_skip_row(name) or _is_skip_row(str(col_a or '')):
            continue

        broker = ws[f'C{row}'].value
        broker = str(broker).strip() if broker else default_broker
        inv_type = ws[f'D{row}'].value or ''
        qty = ws[f'E{row}'].value
        cost = ws[f'F{row}'].value

        # Determine currency
        currency = default_currency
        inv_str = str(inv_type).strip()
        if inv_str in _CURRENCY_MAP:
            currency = _CURRENCY_MAP[inv_str]

        ticker = _resolve_ticker(name, col_a, market, name_map)

        if qty is not None and qty != 0 and cost is not None:
            positions.append({
                'ticker': ticker,
                'name': name,
                'market': market,
                'broker': broker,
                'currency': currency,
                'quantity': float(qty),
                'cost_price': float(cost),
            })

    return positions, closed


def _parse_b_shares(ws, name_map):
    """B股 sheet: two currency groups (USD then HKD), determined by column D."""
    positions = []
    closed = []
    current_currency = 'USD'
    in_closed_section = False

    for row in range(5, ws.max_row + 1):
        col_a = ws[f'A{row}'].value
        name = ws[f'B{row}'].value

        marker = str(col_a or '') + str(name or '')
        if '清仓盈亏明细' in marker:
            in_closed_section = True
            continue

        if in_closed_section:
            trade_name = str(col_a or '').strip() if col_a else ''
            if not trade_name or _is_skip_row(trade_name):
                continue
            pnl = ws[f'J{row}'].value
            if pnl is not None and isinstance(pnl, (int, float)) and pnl != 0:
                closed.append({
                    'ticker': name_map.get(trade_name, trade_name),
                    'name': trade_name,
                    'market': 'B股',
                    'broker': '中信证券',
                    'currency': 'CNY',  # closed P&L in B股 is recorded in RMB
                    'realized_pnl': float(pnl),
                })
            continue

        # Detect currency from column D
        inv_type = ws[f'D{row}'].value
        if inv_type and isinstance(inv_type, str):
            inv_str = inv_type.strip()
            if inv_str in _CURRENCY_MAP:
                current_currency = _CURRENCY_MAP[inv_str]

        if not name or not isinstance(name, str):
            continue
        name = str(name).strip()
        if _is_skip_row(name) or _is_skip_row(str(col_a or '')):
            continue

        broker = ws[f'C{row}'].value
        broker = str(broker).strip() if broker else '中信证券'
        qty = ws[f'E{row}'].value
        cost = ws[f'F{row}'].value
        ticker = _resolve_ticker(name, col_a, 'B股', name_map)

        if qty is not None and qty != 0 and cost is not None:
            positions.append({
                'ticker': ticker,
                'name': name,
                'market': 'B股',
                'broker': broker,
                'currency': current_currency,
                'quantity': float(qty),
                'cost_price': float(cost),
            })

    return positions, closed


def _parse_cash_sheet(ws):
    """Parse 现金 sheet. Row 1=headers (A=账户, B=人民币, C=美元, D=港币)."""
    entries = []
    currencies = {}
    for col in ('B', 'C', 'D'):
        h = ws[f'{col}1'].value
        if h:
            h = str(h).strip()
            if '人民币' in h or 'RMB' in h:
                currencies[col] = 'CNY'
            elif '美元' in h or 'USD' in h:
                currencies[col] = 'USD'
            elif '港币' in h or 'HKD' in h:
                currencies[col] = 'HKD'

    for row in range(2, ws.max_row + 1):
        account = ws[f'A{row}'].value
        if not account or not isinstance(account, str):
            continue
        account = account.strip()
        if '合计' in account:
            continue
        for col, cur in currencies.items():
            val = ws[f'{col}{row}'].value
            if val is not None and isinstance(val, (int, float)):
                entries.append({'account': account, 'currency': cur, 'balance': float(val)})
    return entries


def _parse_nav_history(ws):
    """Parse 盈亏日记账: A=日期, B=净资产, C=投入, D=盈亏, E=权益净值, G=沪深300."""
    entries = []
    for row in range(2, ws.max_row + 1):
        date_val = ws[f'A{row}'].value
        nav = ws[f'B{row}'].value
        if date_val is None or nav is None:
            continue
        date_str = date_val.strftime('%Y-%m-%d') if hasattr(date_val, 'strftime') else str(date_val).strip()
        if not date_str:
            continue
        entries.append({
            'date': date_str,
            'nav': float(nav) if nav else None,
            'capital': float(c) if (c := ws[f'C{row}'].value) else None,
            'pnl': float(p) if (p := ws[f'D{row}'].value) else None,
            'equity_nav': float(e) if (e := ws[f'E{row}'].value) else None,
            'benchmark': float(b) if isinstance(b := ws[f'G{row}'].value, (int, float)) else None,
        })
    return entries


def _parse_fx_rates(ws):
    """Parse FX rates from 资产配置汇总 (C column)."""
    rates = {}
    for row in range(4, ws.max_row + 1):
        cur_name = ws[f'B{row}'].value
        rate = ws[f'C{row}'].value
        if cur_name and rate and isinstance(rate, (int, float)):
            cn = str(cur_name).strip()
            if '日' in cn or 'JPY' in cn:
                rates['JPY'] = float(rate)
            elif '美' in cn or 'USD' in cn:
                rates['USD'] = float(rate)
            elif '港' in cn or 'HKD' in cn:
                rates['HKD'] = float(rate)
    return rates


def _parse_leverage(wb):
    """Parse leverage from 资产配置汇总 sheet.

    Structure:
    - Row 3: headers (A=投资类型, B=币种, C=汇率, D=总资产市值-原币, E=场内杠杆-原币, ...)
    - Rows 4-18: per-market data — column E has 场内杠杆 in original currency
    - Row 23: H=场外杠杆, I=amount (in CNY)
    Returns list of {broker, category, currency, amount (CNY)}.
    """
    results = []
    if '资产配置汇总' not in wb.sheetnames:
        return results
    ws = wb['资产配置汇总']

    # ── 场内杠杆: column E (原币), convert to CNY via column C (汇率) ──
    in_house_total_cny = 0
    for row in range(4, 19):
        leverage_orig = ws[f'E{row}'].value
        rate = ws[f'C{row}'].value
        if leverage_orig and isinstance(leverage_orig, (int, float)) and leverage_orig != 0:
            r = float(rate) if rate and isinstance(rate, (int, float)) else 1.0
            in_house_total_cny += abs(float(leverage_orig)) * r

    if in_house_total_cny > 0:
        results.append({
            'broker': '合计',
            'category': 'in_house',
            'currency': 'CNY',
            'amount': in_house_total_cny,
        })

    # ── 场外杠杆: scan for label in column H, amount in column I ──
    for row in range(19, ws.max_row + 1):
        label = ws[f'H{row}'].value
        if label and isinstance(label, str) and '场外杠杆' in label:
            val = ws[f'I{row}'].value
            if val is not None and isinstance(val, (int, float)) and val != 0:
                results.append({
                    'broker': '合计',
                    'category': 'off_exchange',
                    'currency': 'CNY',
                    'amount': abs(float(val)),
                })
            break

    return results


# ── Main import ───────────────────────────────────────────

def import_all(excel_path=None):
    path = excel_path or EXCEL_PATH
    if not os.path.exists(path):
        print(f"Error: file not found: {path}")
        sys.exit(1)

    print(f"Loading {path} ...")
    wb = openpyxl.load_workbook(path, data_only=True)
    print(f"Sheets: {wb.sheetnames}")

    name_map = _build_name_ticker_map(wb)

    init_db()
    conn = get_conn()

    # Clear for clean import
    for tbl in ('positions', 'closed_trades', 'cash_balances', 'nav_history', 'fx_rates', 'margin_balances'):
        conn.execute(f"DELETE FROM {tbl}")
    conn.commit()

    total_pos = total_closed = 0

    # ── Market sheets ──
    sheet_configs = [
        ('A股',  'A股',  'CNY', '中信证券'),
        ('港股',  '港股',  'HKD', None),
        ('美股',  '美股',  'USD', '富途'),
        ('日股',  '日股',  'JPY', '富途'),
        ('基金',  '基金',  'CNY', '支付宝'),
    ]
    for sheet_name, market, currency, broker in sheet_configs:
        if sheet_name not in wb.sheetnames:
            continue
        pos, closed = _parse_position_sheet(wb[sheet_name], market, currency, broker, name_map)
        for p in pos:
            upsert_position(conn, **p)
        for c in closed:
            insert_closed_trade(conn, **c)
        total_pos += len(pos)
        total_closed += len(closed)
        print(f"  {market}: {len(pos)} positions, {len(closed)} closed")

    # ── B股 (special: dual currency) ──
    if 'B股' in wb.sheetnames:
        pos, closed = _parse_b_shares(wb['B股'], name_map)
        for p in pos:
            upsert_position(conn, **p)
        for c in closed:
            insert_closed_trade(conn, **c)
        total_pos += len(pos)
        total_closed += len(closed)
        print(f"  B股: {len(pos)} positions, {len(closed)} closed")

    # ── 现金 ──
    if '现金' in wb.sheetnames:
        entries = _parse_cash_sheet(wb['现金'])
        for e in entries:
            upsert_cash(conn, **e)
        print(f"  现金: {len(entries)} entries")

    # ── 盈亏日记账 ──
    if '盈亏日记账' in wb.sheetnames:
        entries = _parse_nav_history(wb['盈亏日记账'])
        for e in entries:
            upsert_nav(conn, e['date'], e['nav'], e['capital'], e['pnl'],
                       e['equity_nav'], e['benchmark'])
        print(f"  NAV history: {len(entries)} entries")

    # ── FX rates ──
    if '资产配置汇总' in wb.sheetnames:
        rates = _parse_fx_rates(wb['资产配置汇总'])
        for cur, rate in rates.items():
            upsert_fx(conn, cur, rate)
        print(f"  FX rates: {rates}")

    # ── Leverage / Margin ──
    leverage_entries = _parse_leverage(wb)
    for le in leverage_entries:
        upsert_margin(conn, **le)
    if leverage_entries:
        print(f"  Leverage: {len(leverage_entries)} entries — "
              + ', '.join(f"{e['category']}=¥{e['amount']:,.0f}" for e in leverage_entries))

    conn.commit()
    conn.close()
    print(f"\nDone: {total_pos} positions, {total_closed} closed trades")


if __name__ == '__main__':
    import_all()
